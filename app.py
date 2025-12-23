"""
AI Chat Assistant - Flask Backend
Серверная часть для чат-бота с ИИ
"""

import os
import json
import subprocess
import uuid
from datetime import datetime
from flask import Flask, render_template, request, jsonify, send_file, send_from_directory
from werkzeug.utils import secure_filename

app = Flask(__name__, static_folder='static', template_folder='templates')

# Конфигурация
app.config['SECRET_KEY'] = 'your-secret-key-here'
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['PREVIEW_FOLDER'] = 'previews'
app.config['COMPONENTS_FOLDER'] = 'components'
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB max

# Разрешенные расширения файлов
ALLOWED_DATA_EXTENSIONS = {'csv', 'xlsx', 'xls'}
ALLOWED_PREVIEW_EXTENSIONS = {'png', 'jpg', 'jpeg', 'html'}

# История чата (в production используйте базу данных)
chat_history = []

# Текущая выбранная модель
current_model = {
    'id': None,
    'name': None
}

# Доступные модели для выбора
AVAILABLE_MODELS = {
    'gpt-4': {
        'name': 'GPT-4',
        'description': 'Наиболее мощная модель OpenAI',
        'provider': 'OpenAI'
    },
    'gpt-3.5-turbo': {
        'name': 'GPT-3.5 Turbo',
        'description': 'Быстрая и эффективная модель',
        'provider': 'OpenAI'
    },
    'claude-3': {
        'name': 'Claude 3',
        'description': 'Продвинутая модель Anthropic',
        'provider': 'Anthropic'
    },
    'llama-3': {
        'name': 'LLaMA 3',
        'description': 'Открытая модель от Meta',
        'provider': 'Meta'
    },
    'mistral-7b': {
        'name': 'Mistral 7B',
        'description': 'Компактная и быстрая модель',
        'provider': 'Mistral AI'
    },
    'gemini-pro': {
        'name': 'Gemini Pro',
        'description': 'Мультимодальная модель Google',
        'provider': 'Google'
    }
}

# Доступные компоненты для установки
AVAILABLE_COMPONENTS = {
    'llm-model': {
        'name': 'LLM Модель',
        'script': 'install_llm.py',
        'size': '4.2 GB'
    },
    'embeddings': {
        'name': 'Embeddings',
        'script': 'install_embeddings.py',
        'size': '1.1 GB'
    },
    'voice-synth': {
        'name': 'Голосовой синтез',
        'script': 'install_voice.py',
        'size': '890 MB'
    },
    'image-gen': {
        'name': 'Генерация изображений',
        'script': 'install_imagegen.py',
        'size': '2.3 GB'
    },
    'code-assist': {
        'name': 'Код-ассистент',
        'script': 'install_codeassist.py',
        'size': '1.5 GB'
    },
    'translator': {
        'name': 'Переводчик',
        'script': 'install_translator.py',
        'size': '650 MB'
    }
}


def ensure_directories():
    """Создание необходимых директорий"""
    directories = [
        app.config['UPLOAD_FOLDER'],
        app.config['PREVIEW_FOLDER'],
        app.config['COMPONENTS_FOLDER'],
        'templates',
        'static'
    ]
    for directory in directories:
        os.makedirs(directory, exist_ok=True)


def allowed_file(filename, allowed_extensions):
    """Проверка расширения файла"""
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in allowed_extensions


def generate_ai_response(user_message):
    """
    Генерация ответа ИИ
    В production здесь будет вызов реальной модели
    """
    # Пример простой логики ответа
    lower_message = user_message.lower()
    
    response = {
        'text': '',
        'preview': None
    }
    
    if 'привет' in lower_message or 'здравствуй' in lower_message:
        response['text'] = 'Привет! Рад вас видеть. Чем могу помочь?'
    elif 'картинк' in lower_message or 'изображен' in lower_message or 'фото' in lower_message:
        response['text'] = 'Вот найденные изображения:'
        response['preview'] = {
            'type': 'images',
            'items': get_preview_images()
        }
    elif 'html' in lower_message or 'страниц' in lower_message:
        response['text'] = 'Вот сгенерированный HTML:'
        response['preview'] = {
            'type': 'html',
            'content': generate_sample_html()
        }
    elif 'помощь' in lower_message or 'помоги' in lower_message:
        response['text'] = '''Я могу помочь вам с различными задачами:
• Ответить на вопросы
• Показать изображения (напишите "покажи картинки")
• Сгенерировать HTML (напишите "создай html")
• Обработать загруженные файлы CSV/XLSX
• Установить дополнительные компоненты (нажмите ⚙️)'''
    else:
        responses = [
            'Интересный вопрос! Давайте разберемся вместе.',
            'Хороший вопрос. Вот что я думаю по этому поводу...',
            'Спасибо за вопрос! Позвольте объяснить.',
            'Это увлекательная тема. Расскажу подробнее.',
            'Понимаю вас. Давайте обсудим это детальнее.'
        ]
        import random
        response['text'] = random.choice(responses)
    
    return response


def get_preview_images():
    """Получение списка изображений для превью"""
    images = []
    preview_folder = app.config['PREVIEW_FOLDER']
    
    if os.path.exists(preview_folder):
        for filename in os.listdir(preview_folder):
            if allowed_file(filename, {'png', 'jpg', 'jpeg'}):
                images.append({
                    'name': filename,
                    'url': f'/api/preview/image/{filename}'
                })
    
    # Если нет локальных изображений, используем placeholder
    if not images:
        images = [
            {'name': 'Sample 1', 'url': 'https://picsum.photos/800/600?random=1'},
            {'name': 'Sample 2', 'url': 'https://picsum.photos/800/600?random=2'}
        ]
    
    return images


def generate_sample_html():
    """Генерация примера HTML"""
    return '''<!DOCTYPE html>
<html>
<head>
    <style>
        body { font-family: Arial, sans-serif; padding: 20px; background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); min-height: 100vh; margin: 0; }
        .container { background: white; border-radius: 16px; padding: 30px; max-width: 600px; margin: 0 auto; box-shadow: 0 10px 40px rgba(0,0,0,0.2); }
        h1 { color: #6366f1; margin-bottom: 20px; }
        p { color: #64748b; line-height: 1.6; }
        .badge { display: inline-block; background: #e0e7ff; color: #6366f1; padding: 5px 15px; border-radius: 20px; font-size: 14px; }
    </style>
</head>
<body>
    <div class="container">
        <h1>🚀 Сгенерированная страница</h1>
        <p>Это пример HTML-контента, созданного AI-ассистентом. Вы можете использовать подобные шаблоны для быстрого прототипирования.</p>
        <span class="badge">AI Generated</span>
    </div>
</body>
</html>'''


# ==================== ROUTES ====================

@app.route('/')
def index():
    """Главная страница"""
    return render_template('index.html')


@app.route('/api/check-directory', methods=['GET'])
def check_directory():
    """Проверка наличия директории компонентов"""
    components_path = app.config['COMPONENTS_FOLDER']
    
    # Проверяем существование директории и наличие установленных компонентов
    if not os.path.exists(components_path):
        return jsonify({
            'exists': False,
            'message': 'Директория компонентов не найдена',
            'components': list(AVAILABLE_COMPONENTS.keys())
        })
    
    # Проверяем какие компоненты установлены
    installed = []
    for comp_id in AVAILABLE_COMPONENTS:
        comp_path = os.path.join(components_path, comp_id)
        if os.path.exists(comp_path):
            installed.append(comp_id)
    
    missing = [c for c in AVAILABLE_COMPONENTS if c not in installed]
    
    return jsonify({
        'exists': True,
        'installed': installed,
        'missing': missing,
        'all_installed': len(missing) == 0
    })


@app.route('/api/components', methods=['GET'])
def get_components():
    """Получение списка доступных компонентов"""
    components = []
    for comp_id, comp_data in AVAILABLE_COMPONENTS.items():
        comp_path = os.path.join(app.config['COMPONENTS_FOLDER'], comp_id)
        components.append({
            'id': comp_id,
            'name': comp_data['name'],
            'size': comp_data['size'],
            'installed': os.path.exists(comp_path)
        })
    return jsonify({'components': components})


@app.route('/api/models', methods=['GET'])
def get_models():
    """Получение списка доступных моделей"""
    models = []
    for model_id, model_data in AVAILABLE_MODELS.items():
        models.append({
            'id': model_id,
            'name': model_data['name'],
            'description': model_data['description'],
            'provider': model_data['provider'],
            'selected': current_model['id'] == model_id
        })
    return jsonify({
        'models': models,
        'current': current_model
    })


@app.route('/api/models/select', methods=['POST'])
def select_model():
    """Выбор модели для использования"""
    global current_model
    
    data = request.get_json()
    model_id = data.get('model_id')
    
    if not model_id:
        return jsonify({'success': False, 'error': 'Модель не указана'}), 400
    
    if model_id not in AVAILABLE_MODELS:
        return jsonify({'success': False, 'error': 'Неизвестная модель'}), 400
    
    model_data = AVAILABLE_MODELS[model_id]
    
    # Сохраняем выбранную модель
    current_model = {
        'id': model_id,
        'name': model_data['name'],
        'provider': model_data['provider']
    }
    
    # Сохраняем выбор в файл конфигурации
    config_path = os.path.join(app.config['COMPONENTS_FOLDER'], 'model_config.json')
    os.makedirs(app.config['COMPONENTS_FOLDER'], exist_ok=True)
    
    with open(config_path, 'w') as f:
        json.dump({
            'selected_model': current_model,
            'selected_at': datetime.now().isoformat()
        }, f, indent=2)
    
    return jsonify({
        'success': True,
        'model': current_model,
        'message': f'Модель {model_data["name"]} успешно выбрана'
    })


@app.route('/api/models/current', methods=['GET'])
def get_current_model():
    """Получение текущей выбранной модели"""
    return jsonify({
        'success': True,
        'model': current_model
    })


@app.route('/api/install', methods=['POST'])
def install_components():
    """Установка выбранных компонентов"""
    data = request.get_json()
    components_to_install = data.get('components', [])
    
    if not components_to_install:
        return jsonify({'success': False, 'error': 'Не выбраны компоненты'}), 400
    
    results = []
    for comp_id in components_to_install:
        if comp_id not in AVAILABLE_COMPONENTS:
            results.append({
                'id': comp_id,
                'success': False,
                'error': 'Неизвестный компонент'
            })
            continue
        
        comp_data = AVAILABLE_COMPONENTS[comp_id]
        script_path = os.path.join('scripts', comp_data['script'])
        
        try:
            # Создаем директорию компонента
            comp_path = os.path.join(app.config['COMPONENTS_FOLDER'], comp_id)
            os.makedirs(comp_path, exist_ok=True)
            
            # Запускаем скрипт установки если он существует
            if os.path.exists(script_path):
                result = subprocess.run(
                    ['python', script_path, '--target', comp_path],
                    capture_output=True,
                    text=True,
                    timeout=300  # 5 минут таймаут
                )
                
                if result.returncode != 0:
                    results.append({
                        'id': comp_id,
                        'success': False,
                        'error': result.stderr
                    })
                    continue
            
            # Создаем маркер успешной установки
            marker_path = os.path.join(comp_path, '.installed')
            with open(marker_path, 'w') as f:
                f.write(datetime.now().isoformat())
            
            results.append({
                'id': comp_id,
                'success': True,
                'name': comp_data['name']
            })
            
        except subprocess.TimeoutExpired:
            results.append({
                'id': comp_id,
                'success': False,
                'error': 'Превышено время установки'
            })
        except Exception as e:
            results.append({
                'id': comp_id,
                'success': False,
                'error': str(e)
            })
    
    success_count = sum(1 for r in results if r['success'])
    
    return jsonify({
        'success': success_count > 0,
        'results': results,
        'message': f'Установлено {success_count} из {len(components_to_install)} компонентов'
    })


@app.route('/api/chat', methods=['POST'])
def chat():
    """Обработка сообщения чата"""
    data = request.get_json()
    user_message = data.get('message', '').strip()
    
    if not user_message:
        return jsonify({'success': False, 'error': 'Пустое сообщение'}), 400
    
    # Сохраняем сообщение пользователя
    user_msg = {
        'id': str(uuid.uuid4()),
        'role': 'user',
        'content': user_message,
        'timestamp': datetime.now().isoformat()
    }
    chat_history.append(user_msg)
    
    # Генерируем ответ ИИ
    ai_response = generate_ai_response(user_message)
    
    # Сохраняем ответ ИИ
    ai_msg = {
        'id': str(uuid.uuid4()),
        'role': 'assistant',
        'content': ai_response['text'],
        'preview': ai_response.get('preview'),
        'timestamp': datetime.now().isoformat()
    }
    chat_history.append(ai_msg)
    
    return jsonify({
        'success': True,
        'response': ai_msg
    })


@app.route('/api/chat/history', methods=['GET'])
def get_chat_history():
    """Получение истории чата"""
    return jsonify({
        'success': True,
        'history': chat_history
    })


@app.route('/api/chat/clear', methods=['POST'])
def clear_chat_history():
    """Очистка истории чата"""
    global chat_history
    chat_history = []
    return jsonify({'success': True, 'message': 'История очищена'})


@app.route('/api/upload', methods=['POST'])
def upload_file():
    """Загрузка файлов (CSV, XLSX)"""
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'Файл не найден'}), 400
    
    file = request.files['file']
    
    if file.filename == '':
        return jsonify({'success': False, 'error': 'Файл не выбран'}), 400
    
    if not allowed_file(file.filename, ALLOWED_DATA_EXTENSIONS):
        return jsonify({
            'success': False, 
            'error': f'Неподдерживаемый формат. Разрешены: {", ".join(ALLOWED_DATA_EXTENSIONS)}'
        }), 400
    
    try:
        filename = secure_filename(file.filename)
        # Добавляем уникальный идентификатор к имени файла
        unique_filename = f"{uuid.uuid4().hex[:8]}_{filename}"
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], unique_filename)
        
        file.save(filepath)
        
        # Получаем информацию о файле
        file_info = {
            'id': unique_filename,
            'original_name': filename,
            'size': os.path.getsize(filepath),
            'path': filepath
        }
        
        # Обработка файла в зависимости от типа
        file_ext = filename.rsplit('.', 1)[1].lower()
        
        if file_ext == 'csv':
            file_info['type'] = 'csv'
            file_info['preview'] = process_csv(filepath)
        elif file_ext in ('xlsx', 'xls'):
            file_info['type'] = 'excel'
            file_info['preview'] = process_excel(filepath)
        
        return jsonify({
            'success': True,
            'file': file_info,
            'message': f'Файл {filename} успешно загружен'
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


def process_csv(filepath):
    """Обработка CSV файла"""
    try:
        import csv
        preview_data = {'headers': [], 'rows': [], 'total_rows': 0}
        
        with open(filepath, 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            headers = next(reader, None)
            if headers:
                preview_data['headers'] = headers
            
            rows = []
            for i, row in enumerate(reader):
                if i < 10:  # Первые 10 строк для превью
                    rows.append(row)
                preview_data['total_rows'] = i + 1
            
            preview_data['rows'] = rows
        
        return preview_data
    except Exception as e:
        return {'error': str(e)}


def process_excel(filepath):
    """Обработка Excel файла"""
    try:
        # Попытка использовать openpyxl если установлен
        import importlib.util
        if importlib.util.find_spec('openpyxl'):
            from openpyxl import load_workbook
            
            wb = load_workbook(filepath, read_only=True)
            sheet = wb.active
            
            preview_data = {'headers': [], 'rows': [], 'total_rows': 0, 'sheets': wb.sheetnames}
            
            rows = list(sheet.iter_rows(values_only=True))
            if rows:
                preview_data['headers'] = [str(c) if c else '' for c in rows[0]]
                preview_data['rows'] = [[str(c) if c else '' for c in row] for row in rows[1:11]]
                preview_data['total_rows'] = len(rows) - 1
            
            wb.close()
            return preview_data
        else:
            return {'error': 'openpyxl не установлен. Установите: pip install openpyxl'}
    except Exception as e:
        return {'error': str(e)}


@app.route('/api/preview/image/<filename>')
def get_preview_image(filename):
    """Отдача изображения для превью"""
    return send_from_directory(app.config['PREVIEW_FOLDER'], filename)


@app.route('/api/preview/html/<filename>')
def get_preview_html(filename):
    """Отдача HTML файла для превью"""
    filepath = os.path.join(app.config['PREVIEW_FOLDER'], filename)
    
    if not os.path.exists(filepath):
        return jsonify({'error': 'Файл не найден'}), 404
    
    with open(filepath, 'r', encoding='utf-8') as f:
        content = f.read()
    
    return jsonify({
        'success': True,
        'content': content,
        'filename': filename
    })


@app.route('/api/preview/generate', methods=['POST'])
def generate_preview():
    """Генерация превью для отображения в боковой панели"""
    data = request.get_json()
    preview_type = data.get('type', 'html')
    
    if preview_type == 'html':
        content = data.get('content') or generate_sample_html()
        return jsonify({
            'success': True,
            'type': 'html',
            'content': content
        })
    
    elif preview_type == 'images':
        images = get_preview_images()
        return jsonify({
            'success': True,
            'type': 'images',
            'items': images
        })
    
    return jsonify({'success': False, 'error': 'Неизвестный тип превью'}), 400


@app.route('/api/files/<path:filename>')
def serve_file(filename):
    """Отдача загруженных файлов"""
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename)


# ==================== ERROR HANDLERS ====================

@app.errorhandler(404)
def not_found(error):
    return jsonify({'error': 'Не найдено'}), 404


@app.errorhandler(500)
def server_error(error):
    return jsonify({'error': 'Внутренняя ошибка сервера'}), 500


@app.errorhandler(413)
def too_large(error):
    return jsonify({'error': 'Файл слишком большой'}), 413


# ==================== MAIN ====================

if __name__ == '__main__':
    ensure_directories()
    print("🚀 AI Chat Assistant запущен!")
    print("📍 Откройте http://localhost:5000 в браузере")
    app.run(debug=True, host='0.0.0.0', port=5000)
